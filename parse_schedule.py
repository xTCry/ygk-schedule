import json
import re
import pandas as pd
from openpyxl import load_workbook

# file_path = 'so__1_.xlsx'
input_file_path = 'files/2025-2/so__2.xlsx'
output_file_path = 'files/2025-2/schedule.json'

days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

wb = load_workbook(filename=input_file_path, data_only=True)
print('[wb] sheetnames', wb.sheetnames)

xls = pd.ExcelFile(input_file_path)
print('[xls] sheet_names', xls.sheet_names)

all_schedules = {}

SHOW_DEBUG = False

for sheet_name in xls.sheet_names:
    ws = wb[sheet_name]
    df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=None)
    # print(df)

    schedule = {}
    current_day = None
    current_group = None

    prev_lesson_row = None
    prev_lesson_number = -1

    counter_test = 0

    # парсинг листа с группами
    for rowIndex, row in df.iterrows():
        first_cell = str(row[0]).strip()
        other_cells_are_nan = all([pd.isna(cell) for cell in row[1:]])

        if first_cell.startswith('* с ДОТ'):
            current_day = None
            current_group = None
            prev_lesson_row = None
            prev_lesson_number = -1
            continue

        # debug vars
        if SHOW_DEBUG and (not pd.isna(row[0]) or not other_cells_are_nan):
            print('\n=========== ', '[', rowIndex+1, ']', sheet_name)
            print('other & first_cell:', other_cells_are_nan, first_cell)
            print('current_group:', current_group)
            print('current_day:', current_day)
            print('second_lesson_number:', prev_lesson_number)
            print('^^^^^^^^^^^^^^^^^')

        # определяем день недели
        if other_cells_are_nan and current_group and first_cell in days:
            current_day = first_cell
            if current_day not in schedule[current_group]:
                schedule[current_group][current_day] = []
            prev_lesson_number = -1
            if SHOW_DEBUG:
                print('[detected] current_day:', current_day)
            continue

        # определяем название группы
        if other_cells_are_nan and not pd.isna(row[0]) and len(first_cell) > 3 and len(first_cell) < 32 and '-' in first_cell:
            # TODO: improve regexp
            current_groups = re.findall(r'[A-я0-9\-\/]+', first_cell)
            if SHOW_DEBUG:
                print('[detected] Z current_group:', current_groups)
            if not pd.isna(current_groups[0]):
                current_group = current_groups[0]
                if SHOW_DEBUG:
                    print('[detected] current_group:', current_group)
                if current_group not in schedule:
                    schedule[current_group] = {}
                current_day = None
                continue

        # определяем номер пары (и проверка на двойную ячейку подгруппы)
        if current_day and (first_cell.isdigit() or prev_lesson_number > -1):
            if not other_cells_are_nan:
                lesson_num = prev_lesson_number if prev_lesson_number > -1 else int(first_cell)

                discipline = str(row[1]).strip() if not pd.isna(row[1]) else ''
                if discipline == '' and not pd.isna(prev_lesson_row[1]) and prev_lesson_number > -1:
                    discipline = str(prev_lesson_row[1]).strip() if not pd.isna(prev_lesson_row[1]) else ''

                teacher = str(row[5]).strip() if not pd.isna(row[5]) else ''
                if teacher == '' and not pd.isna(prev_lesson_row[5]) and prev_lesson_number > -1:
                    teacher = str(prev_lesson_row[5]).strip() if not pd.isna(prev_lesson_row[5]) else ''

                auditory = str(row[8]).strip() if not pd.isna(row[8]) else ''
                if auditory == '' and not pd.isna(prev_lesson_row[8]) and prev_lesson_number > -1:
                    auditory = str(prev_lesson_row[8]).strip() if not pd.isna(prev_lesson_row[8]) else ''

                isRedDiscipline = False
                isRedAuditory = False
                colorDiscipline = ws.cell(row=rowIndex+1, column=2).fill.fgColor.rgb
                if colorDiscipline != '00000000' and str(colorDiscipline).lower() != 'ffffffff':
                    isRedDiscipline = True
                    print('[color] colorDiscipline', rowIndex+1, colorDiscipline)

                colorAuditory = ws.cell(row=rowIndex+1, column=9).fill.fgColor.rgb
                if colorAuditory != '00000000' and str(colorAuditory).lower() != 'ffffffff':
                    isRedAuditory = True
                    print('[color] colorAuditory', rowIndex+1, colorAuditory)

                if SHOW_DEBUG:
                    print('[detected] lesson:', lesson_num, '.', discipline, '[', teacher, ']', auditory)

                entry = {
                    'номер': lesson_num,
                    'дисциплина': discipline,
                    'преподаватель': teacher,
                    'аудитория': auditory,
                    'красный-дисциплина': isRedDiscipline,
                    'красный-аудитория': isRedAuditory,
                }
                schedule[current_group][current_day].append(entry)
            
            # reset
            if prev_lesson_number > -1:
                prev_lesson_number = -1
            else:
                prev_lesson_number = int(first_cell)
            prev_lesson_row = row.copy()

        counter_test += 1
        # if counter_test > 29:
        #     break

    all_schedules[sheet_name] = schedule

    # # ! for test
    # break

with open(output_file_path, 'w', encoding='utf-8') as f:
    json.dump(all_schedules, f, ensure_ascii=False, indent=2)

print('Saved', all_schedules.keys())
